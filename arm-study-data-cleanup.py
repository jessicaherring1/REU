### ORGANIZE EXCEL SHEET FOR QUALTRICS ARM MORPHOLOGY STUDY ###

from openpyxl import Workbook, load_workbook
import math

# old workbook varibales 
oWorkbook = load_workbook(filename="Documents/REU/old-arm-study-data.xlsx")
oSheet = oWorkbook.active

# updated workbook (formatted how i like) 
uWorkbook = Workbook()
uSheet = uWorkbook.active

#averaged-questions workbook (updated from uWorkbook, averages individual participant answers from question groups) 
qWorkbook = Workbook()
qSheet = qWorkbook.active

#averaged workbook (updated from uWorkbook, averages all participant answers per row) 
aWorkbook = Workbook()
aSheet = aWorkbook.active

#simplified workbook (updated from aWorkbook, simplifies questions into their categories (competency, discomfort, etc.)) 
sWorkbook = Workbook()
sSheet = sWorkbook.active


def main():

    uSheetFunc()

    qSheetFunc()

    aSheetFunc()

    sSheetFunc()

    csvFunc()



## UPDATES THE ORIGINAL EXCEL SHEET BY CONVERTING ALL COLUMNS INTO ROWS AND DELETE UNNECESSARY DATA
def uSheetFunc():

    # adds "blank"/title row to uSheet help with pandas processing
    uSheet["A1"] = "1Title"

    ## CONVERT (columns into rows)

    for i in range(1, oSheet.max_column):
        names = []

        #collect all information in column into an array
        for row in oSheet:
            name = row[i].value
            names.append(name)

        #add array into row
        for b in range(len(names)):
            temp = names[b]
            uSheet.cell(row=i+1, column=b+1).value = temp


    ## CLEAN OUT (delete rows that deal with not-arm specific data, timing, free-response answers, and price senstivity questions)

    # outer for-loop ensures no lines are skipped as lines are deleted (very inefficient lol)
    for i in range(1, uSheet.max_row):

        #for each row in the sheet
        for row in uSheet:

            #value in first column
            string = str(row[0].value) 

            # first letter of value in first column
            first_letter = string[0]

            #if first letter is not a number, delete the row
            if not first_letter.isnumeric():
                uSheet.delete_rows(row[0].row, 1)  

            #if stim, ps, or q49 are substrings within the cell value, delete the row
            if "stim" in string or "PS" in string or "Q49" in string:
                uSheet.delete_rows(row[0].row, 1)   



    #delete columns B and C because they contain unnecessary data
    uSheet.delete_cols(2, 2) 

    # save into new excel file
    uWorkbook.save(filename="Documents/REU/updated-arm-study-data.xlsx")


def qSheetFunc():

    #creates new titles for first column
    #   pattern: armNumber_characteristic
    for i in range(1, 28):
        if i % 3 == 2:
            insert("A", i+1, str(math.ceil(i/3)) + "_Competency", qSheet ) 
        elif i % 3 == 1:
            insert("A", i+1, str(math.ceil(i/3)) + "_Discomfort", qSheet ) 
        elif i % 3 == 0:
            insert("A", i+1, str(math.ceil(i/3)) + "_Safety", qSheet ) 


    #compiles question averages for each person (all columns except title)
    #i can be used for column number
    for i in range(2, (uSheet.max_column + 1)):

        #gets discomfort, competency, an safety averages for each arm
        comp1, disc1, saf1 = simplify2(1, i)
        comp2, disc2, saf2 = simplify2(2, i)
        comp3, disc3, saf3 = simplify2(3, i)
        comp4, disc4, saf4 = simplify2(4, i)
        comp5, disc5, saf5 = simplify2(5, i)
        comp6, disc6, saf6 = simplify2(6, i)
        comp7, disc7, saf7 = simplify2(7, i)
        comp8, disc8, saf8 = simplify2(8, i)
        comp9, disc9, saf9 = simplify2(9, i)

        #places averages into designated spot on qSheet
        qSheet.cell(row=2, column=i).value = comp1
        qSheet.cell(row=3, column=i).value = disc1
        qSheet.cell(row=4, column=i).value = saf1
        qSheet.cell(row=5, column=i).value = comp2
        qSheet.cell(row=6, column=i).value = disc2
        qSheet.cell(row=7, column=i).value = saf2
        qSheet.cell(row=8, column=i).value = comp3
        qSheet.cell(row=9, column=i).value = disc3
        qSheet.cell(row=10, column=i).value = saf3
        qSheet.cell(row=11, column=i).value = comp4
        qSheet.cell(row=12, column=i).value = disc4
        qSheet.cell(row=13, column=i).value = saf4
        qSheet.cell(row=14, column=i).value = comp5
        qSheet.cell(row=15, column=i).value = disc5
        qSheet.cell(row=16, column=i).value = saf5
        qSheet.cell(row=17, column=i).value = comp6
        qSheet.cell(row=18, column=i).value = disc6
        qSheet.cell(row=19, column=i).value = saf6
        qSheet.cell(row=20, column=i).value = comp7
        qSheet.cell(row=21, column=i).value = disc7
        qSheet.cell(row=22, column=i).value = saf7
        qSheet.cell(row=23, column=i).value = comp8
        qSheet.cell(row=24, column=i).value = disc8
        qSheet.cell(row=25, column=i).value = saf8
        qSheet.cell(row=26, column=i).value = comp9
        qSheet.cell(row=27, column=i).value = disc9
        qSheet.cell(row=28, column=i).value = saf9

        #adds column headers
        insert("A", 1, "Title", qSheet ) 
        insert("B", 1, "Average", qSheet ) 

    # save into new excel file
    qWorkbook.save(filename="Documents/REU/averaged-questions-arm-study-data.xlsx")

## AVERAGED WORKBOOK -- AVERAGES ALL PARTIPANT VALUES AND PUTS INTO ONE COLUMN
def aSheetFunc():
    
    num_col = uSheet.max_column

    #copy over names from uSheet first column to aSheet first column
    for i in range (1, (uSheet.max_row + 1)):
        rows = uSheet.iter_cols(min_row=i, max_col=1, values_only=True)
        
        #all titles 
        values = [row[0] for row in rows]

        location = "A" + str(i)
        
        aSheet[location] = values[0]


    #averaged all participant responses and copied them into second column
    for i in range (2, (uSheet.max_row + 1)):
        rows = uSheet.iter_cols(min_row=i, min_col=2, max_col=num_col, values_only=True)
        
        #all participant values 
        values = [row[0] for row in rows]

        avg = sum(values) / len(values)

        location = "B" + str(i)

        aSheet[location] = avg

    #column title
    aSheet["B1"] = "Average"

    # save into new excel file
    aWorkbook.save(filename="Documents/REU/averaged-arm-study-data.xlsx")



## SIMPLIFIED WORKBOOK -- COMBINES QUESTION GROUPS INTO SINGLE ROW MY AVERAGING DATA (Q1-6 collapsed to "competency", Q7-12 to "discomfort", Q1-3 to "safety")
##          THEN ADDED ALL VALUES TO NEW FILE
def sSheetFunc():
    
    #gets discomfort, competency, an safety averages for each arm
    disc1, comp1, saf1 = simplify("1")
    disc2, comp2, saf2 = simplify("2")
    disc3, comp3, saf3 = simplify("3")
    disc4, comp4, saf4 = simplify("4")
    disc5, comp5, saf5 = simplify("5")
    disc6, comp6, saf6 = simplify("6")
    disc7, comp7, saf7 = simplify("7")
    disc8, comp8, saf8 = simplify("8")
    disc9, comp9, saf9 = simplify("9")

    #creates new titles for first column
    #   pattern: armNumber_characteristic
    for i in range(1, 28):
        if i % 3 == 2:
            insert("A", i+1, str(math.ceil(i/3)) + "_Competency", sSheet ) 
        elif i % 3 == 1:
            insert("A", i+1, str(math.ceil(i/3)) + "_Discomfort", sSheet ) 
        elif i % 3 == 0:
            insert("A", i+1, str(math.ceil(i/3)) + "_Safety", sSheet ) 

    #adds column headers
    insert("A", 1, "Title", sSheet ) 
    insert("B", 1, "Average", sSheet ) 

    #places averages into designated spot on sSheet
    insert("B", 2, disc1, sSheet )
    insert("B", 3, comp1, sSheet ) 
    insert("B", 4, saf1, sSheet ) 
    insert("B", 5, disc2, sSheet ) 
    insert("B", 6, comp2, sSheet ) 
    insert("B", 7, saf2, sSheet ) 
    insert("B", 8, disc3, sSheet ) 
    insert("B", 9, comp3, sSheet ) 
    insert("B", 10, saf3, sSheet ) 
    insert("B", 11, disc4, sSheet ) 
    insert("B", 12, comp4, sSheet )
    insert("B", 13, saf4, sSheet ) 
    insert("B", 14, disc5, sSheet ) 
    insert("B", 15, comp5, sSheet ) 
    insert("B", 16, saf5, sSheet ) 
    insert("B", 17, disc6, sSheet ) 
    insert("B", 18, comp6, sSheet ) 
    insert("B", 19, saf6, sSheet )
    insert("B", 20, disc7, sSheet )
    insert("B", 21, comp7, sSheet ) 
    insert("B", 22, saf7, sSheet ) 
    insert("B", 23, disc8, sSheet ) 
    insert("B", 24, comp8, sSheet ) 
    insert("B", 25, saf8, sSheet )  
    insert("B", 26, disc9, sSheet ) 
    insert("B", 27, comp9, sSheet ) 
    insert("B", 28, saf9, sSheet )  


    # save into new excel file
    sWorkbook.save(filename="Documents/REU/simplified-arm-study-data.xlsx")



#SIMPLIFIES THE QUESTION GROUPS AND RETURNS AVERAGE FOR EACH GROUP
#   num is a string (bad naming haha) and represents the arm number
def simplify(num):

    #question numbers distinguish which characteristic category
    #   1-6: Competency
    #   7-12: Discomfort
    comp = ["1", "2", "3", "4", "5", "6"]
    disc = ["7", "8", "9", "10", "11", "12"]

    #accumulator variables for each characteristic
    disc1 = 0
    comp1 = 0
    saf1 = 0

    #loops for number of rows
    for i in range (1, (aSheet.max_row + 1)):
        rows = aSheet.iter_cols(min_row=i, min_col=1, max_col=2, values_only=True)

        #values in each column of the row
        values = [row[0] for row in rows]

        name = values[0]

        #number of the arm 
        first_letter = str(name)[0]

        #question number
        last_2letter = str(name)[::-1][0:2][::-1]

        #filters for only questions relating to specific arm given as a parameter
        if num in first_letter:

            #rating in cell
            value = values[1]

            if "comp" in name: 
                #discomfort questions 
                if last_2letter.isnumeric() or any([x in last_2letter for x in disc]):
                    disc1 += value
                #competency questions 
                else:
                    comp1 += value
                    
            #safety questions
            elif "Average" not in str(value): 
                saf1 += value
                
    #compute averages
    discomfort = disc1/6
    competency = comp1/6
    safety = saf1/3

    return(discomfort, competency, safety)


#VARIATION OF THE SIMPLIFY FUNCTION ABOVE
#   num is a int that represents the arm number 
#   col is a number that respresents the column number (one person's answers)
def simplify2(num, col):

    #question numbers distinguish which characteristic category
    #   1-6: Competency
    #   7-12: Discomfort
    comp = ["1", "2", "3", "4", "5", "6"]
    disc = ["7", "8", "9", "10", "11", "12"]

    #accumulator variables for each characteristic
    disc1 = 0
    comp1 = 0
    saf1 = 0


    for i in range (2, (uSheet.max_row + 1)):
        rows = uSheet.iter_cols(min_row=i, max_col=col, values_only=True)

        #values in each column of the row
        values = [row[0] for row in rows]

        name = values[0]

        #number of the arm 
        first_letter = str(name)[0]

        #question number
        last_2letter = name[::-1][0:2][::-1]

        if str(num) in first_letter:
            value = values[col-1]

            if "comp" in name: 
                #discomfort questions 
                if last_2letter.isnumeric() or any([x in last_2letter for x in disc]):
                    disc1 += value
                #competency questions 
                else:
                    comp1 += value
            #safety questions
            elif "Average" not in str(value): 
                saf1 += value
                

    #compute averages
    discomfort = disc1/6
    competency = comp1/6
    safety = saf1/3

    return(discomfort, competency, safety)




#INSERTS INFO. INTO SPECIFIED LOCATION IN SSHEET
def insert(col, row, val, sheet):
    loc = col + str(row)
    sheet[loc] = val



#SAVES EXCEL FILES IN CSV FORMAT
def csvFunc():
    ## CREATING CSV FILES
    #https://www.studytonight.com/post/converting-xlsx-file-to-csv-file-using-python#:~:text=You%20can%20use%20openpyxl%20to,standard%20file%20I%2FO%20operations.
    uCsv = open("Documents/REU/updated-arm-study-data.csv", "w+")
    qCsv = open("Documents/REU/averaged-questions-arm-study-data.csv", "w+")
    aCsv = open("Documents/REU/averaged-arm-study-data.csv", "w+")
    sCsv = open("Documents/REU/simplified-arm-study-data.csv", "w+")

    #for updated sheet
    for row in uSheet.rows:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                uCsv.write(str(l[i].value))
                uCsv.write('\n')
            else:
                uCsv.write(str(l[i].value) + ',')
   
    #for averaged-questions sheet
    for row in qSheet.rows:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                qCsv.write(str(l[i].value))
                qCsv.write('\n')
            else:
                qCsv.write(str(l[i].value) + ',')


    #for averaged sheet
    for row in aSheet.rows:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                aCsv.write(str(l[i].value))
                aCsv.write('\n')
            else:
                aCsv.write(str(l[i].value) + ',')

    #for simplified sheet
    for row in sSheet.rows:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                sCsv.write(str(l[i].value))
                sCsv.write('\n')
            else:
                sCsv.write(str(l[i].value) + ',')
            
    uCsv.close()
    qCsv.close()
    aCsv.close()
    sCsv.close()


main()




'''
### NOTES ###

#ADD INDIVIDUAL VALUES TO CELLS 
uSheet["A1"] = "hello"
uSheet["B1"] = "world"

#STORE VALUE OF COLUMN INTO ARRAY
names = []
for row in oSheet:
    name = row[30].value
    names.append(name)

#STORE VALUE OF ARRAY INTO ROW
for b in range(len(names)):
        temp = names[b]
        uSheet.cell(row=1, column=b+1).value = temp


'''