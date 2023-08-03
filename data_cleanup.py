### ORGANIZE EXCEL SHEET FOR QUALTRICS ARM MORPHOLOGY STUDY ###


from openpyxl import Workbook, load_workbook
import math



# workbook that Qualtrics outputs  
oWorkbook = load_workbook(filename="excel-sheets/old-survey-data.xlsx")
oSheet = oWorkbook.active

# duplicated oWorkbook
uWorkbook2 = oWorkbook
uSheet2 = uWorkbook2.active

# updated workbook 
#    updated from oWorkbook, all columns are turned into rows and unnecessary data is deleted 
#    only deals with warmth, discomfort, competency, and safety
uWorkbook = Workbook()
uSheet = uWorkbook.active

# updates workbook 2
#    updated from oWorkbook, all columns are turned into rows and unnecessary data is deleted 
#    only deals with price sensitivity
uWorkbook2 = Workbook()
uSheet2 = uWorkbook2.active

#averaged-questions workbook 
#    updated from uWorkbook, averages individual participant answers within each question group
qWorkbook = Workbook()
qSheet = qWorkbook.active

#averaged workbook 
#    updated from uWorkbook, averages all participant answers per row
aWorkbook = Workbook()
aSheet = aWorkbook.active

#simplified workbook 
#    updated from aWorkbook, simplifies averaged answers into their categories (competency, discomfort, etc.)
sWorkbook = Workbook()
sSheet = sWorkbook.active

#price workbook 
#    updated from oWorkbook2, handles price sensitivity questions
pWorkbook = Workbook()
pSheet = pWorkbook.active

#price 2 workbook 
#    updated from pWorkbook, shifts all rows into columns (for matplotlib box plots)
pWorkbook2 = Workbook()
pSheet2 = pWorkbook.active

#averaged-price workbook 
#    updated from oWorkbook2, handles price sensitivity questions
apWorkbook = Workbook()
apSheet = apWorkbook.active


def main():

    eSheetFunc()

    uSheetFunc()

    uSheetFunc2()

    qSheetFunc()

    aSheetFunc()

    sSheetFunc()

    pSheetFunc()

    apSheetFunc()

    csvFunc()


# DELETES ALL PARTICIPANTS THAT DID NOT COMPLETE TO 100%
def eSheetFunc():
    # outer for-loop ensures no lines are skipped as lines are deleted (very inefficient lol sorry)
    for i in range(1, oSheet.max_row):

        #for each row in the sheet
        for row in oSheet:

            #value in fifth column (progress column)
            progress = row[4].value


            #if progress is not 100 (participant did not finish entire survey), delete the row (first part of if statement)
            #    keeps first three rows bc they contain important info. (second part of if statement)
            if (not progress == 100) and (not "rogress" in str(progress)):
                oSheet.delete_rows(row[0].row, 1)  



## UPDATES THE ORIGINAL EXCEL SHEET BY CONVERTING ALL COLUMNS INTO ROWS AND ONLY KEEPING PRICE SENSITIVITY DATA
def uSheetFunc():
    # adds "blank"/title row to uSheet help with pandas processing
    uSheet["A1"] = "1Title"

    ## convert columns into rows
    for i in range(1, oSheet.max_column):
        names = []

        #collect all information in column into an array
        for row in oSheet:
            name = row[i].value
            names.append(name)

        #add array into row
        for b in range(len(names)):
            temp = names[b]
            uSheet.cell(row=i+1, column=b+1).value = temp


    ## CLEAN OUT (delete rows that deal with not-arm specific data, timing, free-response answers, and price senstivity questions)

    # outer for-loop ensures no lines are skipped as lines are deleted (very inefficient lol sorry)
    for i in range(1, uSheet.max_row):

        #for each row in the sheet
        for row in uSheet:

            #value in first column
            string = str(row[0].value) 

            # first letter of value in first column
            first_letter = string[0]

            #if first letter is not a number, delete the row (deletes non-characteristic questions-- ex. timing, demographics, etc.)
            if not first_letter.isnumeric():
                uSheet.delete_rows(row[0].row, 1)  

            #if stim, ps, or q49 are substrings within the cell value, delete the row
            #   certain questions weren't getting deleted so this handles that 
            #   deletes price sensitivity question too
            if "stim" in string or "PS" in string or "Q49" in string:
                uSheet.delete_rows(row[0].row, 1)   


    #delete columns B and C because they contain unnecessary data
    uSheet.delete_cols(2, 2) 

    # save into new excel file
    uWorkbook.save(filename="excel-sheets/updated-arm-study-data.xlsx")



## UPDATES THE ORIGINAL EXCEL SHEET BY CONVERTING ALL COLUMNS INTO ROWS AND DELETE UNNECESSARY DATA
def uSheetFunc2():

    # adds "blank"/title row to uSheet help with pandas processing
    uSheet2["A1"] = "1Title"

    ## convert columns into rows
    for i in range(1, oSheet.max_column):
        names = []

        #collect all information in column into an array
        for row in oSheet:
            name = row[i].value
            names.append(name)

        #add array into row
        for b in range(len(names)):
            temp = names[b]
            uSheet2.cell(row=i+1, column=b+1).value = temp


    ## CLEAN OUT (delete rows that deal with not-priceSensitivity specific data( ex. timing, free-response answers, and rosas/godspeed questions)

    # outer for-loop ensures no lines are skipped as lines are deleted (very inefficient lol sorry)
    for i in range(1, uSheet2.max_row):

        #for each row in the sheet
        for row in uSheet2:

            #value in first column
            string = str(row[0].value) 

            # first letter of value in first column
            first_letter = string[0]

            #if first letter is not a number, delete the row (deletes non-characteristic questions-- ex. timing, demographics, etc.)
            if not first_letter.isnumeric() and not "Q53" in string:
                uSheet2.delete_rows(row[0].row, 1)  

            #if stim, competency, discomfort, warmth, safety, or q49 are substrings within the cell value, delete the row
            if ("stim" in string) or ("comp" in string) or ("disc" in string) or ("warm" in string) or ("Saf" in string) or ("Q49" in string):
                uSheet2.delete_rows(row[0].row, 1)   


    #delete columns B and C because they contain unnecessary data
    uSheet2.delete_cols(2, 2) 

    # save into new excel file
    uWorkbook2.save(filename="excel-sheets/updated-arm-study-data.xlsx")



#AVERAGES QUESTIONS WORKBOOK 
#   UPDATED FROM UWORKBOOK, AVERAGES INDIVIDUAL PARTICIPANT ANSWERS WITHIN EACH QUESTION GROUP
def qSheetFunc():

    #creates new titles for first column
    #   pattern: armNumber_characteristic (ex. 1_Warmth)
    for i in range(1, 37):
        if i % 4 == 3:
            insert("A", i+1, str(math.ceil(i/4)) + "_Competency", qSheet ) 
        elif i % 4 == 2:
            insert("A", i+1, str(math.ceil(i/4)) + "_Discomfort", qSheet ) 
        elif i % 4 == 1:
            insert("A", i+1, str(math.ceil(i/4)) + "_Warmth", qSheet ) 
        elif i % 4 == 0:
            insert("A", i+1, str(math.ceil(i/4)) + "_Safety", qSheet ) 


    #compiles question averages for each person (all columns except title)
    #i can be used for column number
    print(uSheet.max_column)
    for i in range(2, (uSheet.max_column + 1)):

        #gets discomfort, competency, an safety averages for each arm
        warm1, comp1, disc1, saf1 = simplify2(1, i, uSheet)
        warm2, comp2, disc2, saf2 = simplify2(2, i, uSheet)
        warm3, comp3, disc3, saf3 = simplify2(3, i, uSheet)
        warm4, comp4, disc4, saf4 = simplify2(4, i, uSheet)
        warm5, comp5, disc5, saf5 = simplify2(5, i, uSheet)
        warm6, comp6, disc6, saf6 = simplify2(6, i, uSheet)
        warm7, comp7, disc7, saf7 = simplify2(7, i, uSheet)
        warm8, comp8, disc8, saf8 = simplify2(8, i, uSheet)
        warm9, comp9, disc9, saf9 = simplify2(9, i, uSheet)

        #places averages into designated spot on qSheet
        #   respectfully, this is disgusting
        qSheet.cell(row=2, column=i).value = warm1
        qSheet.cell(row=3, column=i).value = comp1
        qSheet.cell(row=4, column=i).value = disc1
        qSheet.cell(row=5, column=i).value = saf1
        qSheet.cell(row=6, column=i).value = warm2
        qSheet.cell(row=7, column=i).value = comp2
        qSheet.cell(row=8, column=i).value = disc2
        qSheet.cell(row=9, column=i).value = saf2
        qSheet.cell(row=10, column=i).value = warm3
        qSheet.cell(row=11, column=i).value = comp3
        qSheet.cell(row=12, column=i).value = disc3
        qSheet.cell(row=13, column=i).value = saf3
        qSheet.cell(row=14, column=i).value = warm4
        qSheet.cell(row=15, column=i).value = comp4
        qSheet.cell(row=16, column=i).value = disc4
        qSheet.cell(row=17, column=i).value = saf4
        qSheet.cell(row=18, column=i).value = warm5
        qSheet.cell(row=19, column=i).value = comp5
        qSheet.cell(row=20, column=i).value = disc5
        qSheet.cell(row=21, column=i).value = saf5
        qSheet.cell(row=22, column=i).value = warm6
        qSheet.cell(row=23, column=i).value = comp6
        qSheet.cell(row=24, column=i).value = disc6
        qSheet.cell(row=25, column=i).value = saf6
        qSheet.cell(row=26, column=i).value = warm7
        qSheet.cell(row=27, column=i).value = comp7
        qSheet.cell(row=28, column=i).value = disc7
        qSheet.cell(row=29, column=i).value = saf7
        qSheet.cell(row=30, column=i).value = warm8
        qSheet.cell(row=31, column=i).value = comp8
        qSheet.cell(row=32, column=i).value = disc8
        qSheet.cell(row=33, column=i).value = saf8
        qSheet.cell(row=34, column=i).value = warm9
        qSheet.cell(row=35, column=i).value = comp9
        qSheet.cell(row=36, column=i).value = disc9
        qSheet.cell(row=37, column=i).value = saf9

        #adds column headers (helps when working with pandas)
        insert("A", 1, "Title", qSheet ) 
        insert("B", 1, "Average", qSheet ) 

    # save into new excel file
    qWorkbook.save(filename="excel-sheets/averaged-questions-arm-study-data.xlsx")



## AVERAGED WORKBOOK -- AVERAGES ALL PARTIPANT VALUES AND PUTS INTO ONE COLUMN
def aSheetFunc():
    
    #number of columns
    #   represents number of participants 
    num_col = uSheet.max_column

    #copy over names from uSheet first column to aSheet first column
    for i in range (1, (uSheet.max_row + 1)):
        rows = uSheet.iter_cols(min_row=i, max_col=1, values_only=True)
        
        #all titles 
        values = [row[0] for row in rows]

        location = "A" + str(i)
        
        aSheet[location] = values[0]


    #averaged all participant responses and copy them into second column
    for i in range (2, (uSheet.max_row + 1)):
        rows = uSheet.iter_cols(min_row=i, min_col=2, max_col=num_col, values_only=True)
        
        #all participant values 
        values = [row[0] for row in rows]

        newValues = []

        #gets rid of 'None' values that prevent doing arithmetic (came as a result of incomplete qualtrics data, shouldn't be a problem anymore idk)
        for j in range(len(values)):
            if isinstance(values[j] , int):
                newValues.append(values[j])

        avg = sum(newValues) / len(newValues)

        location = "B" + str(i)

        aSheet[location] = avg

    #column title
    aSheet["B1"] = "Average"

    # save into new excel file
    aWorkbook.save(filename="excel-sheets/averaged-arm-study-data.xlsx")



## SIMPLIFIED WORKBOOK -- COMBINES QUESTION GROUPS INTO SINGLE ROW MY AVERAGING DATA (Q1-6 collapsed to "warmth", Q7-12 to "competency", Q13-18 to "discomfort", Q1-3 to "safety")
##          THEN ADDED ALL VALUES TO NEW FILE
def sSheetFunc():
    
    #gets discomfort, competency, an safety averages for each arm
    warm1, disc1, comp1, saf1 = simplify("1")
    warm2, disc2, comp2, saf2 = simplify("2")
    warm3, disc3, comp3, saf3 = simplify("3")
    warm4, disc4, comp4, saf4 = simplify("4")
    warm5, disc5, comp5, saf5 = simplify("5")
    warm6, disc6, comp6, saf6 = simplify("6")
    warm7, disc7, comp7, saf7 = simplify("7")
    warm8, disc8, comp8, saf8 = simplify("8")
    warm9, disc9, comp9, saf9 = simplify("9")

    #creates new titles for first column
    #   pattern: armNumber_characteristic (ex. 1_Competency, 4_Safety, etc.)
    for i in range(1, 37):
        if i % 4 == 3:
            insert("A", i+1, str(math.ceil(i/4)) + "_Competency", sSheet ) 
        elif i % 4 == 2:
            insert("A", i+1, str(math.ceil(i/4)) + "_Discomfort", sSheet ) 
        elif i % 4 == 1:
            insert("A", i+1, str(math.ceil(i/4)) + "_Warmth", sSheet ) 
        elif i % 4 == 0:
            insert("A", i+1, str(math.ceil(i/4)) + "_Safety", sSheet ) 

    #adds column headers (helps w/ pandas/matplotlib)
    insert("A", 1, "Title", sSheet ) 
    insert("B", 1, "Average", sSheet ) 

    #places averages into designated spot on sSheet
    #   sorry this gross, follows a simple pattern though 
    insert("B",  2, warm1, sSheet )
    insert("B",  3, disc1, sSheet )
    insert("B",  4, comp1, sSheet ) 
    insert("B",  5, saf1,  sSheet ) 

    insert("B",  6, warm2, sSheet )
    insert("B",  7, disc2, sSheet ) 
    insert("B",  8, comp2, sSheet ) 
    insert("B",  9, saf2,  sSheet ) 

    insert("B", 10, warm3, sSheet )
    insert("B", 11, disc3, sSheet ) 
    insert("B", 12, comp3, sSheet ) 
    insert("B", 13, saf3,  sSheet ) 

    insert("B", 14, warm4, sSheet )
    insert("B", 15, disc4, sSheet ) 
    insert("B", 16, comp4, sSheet )
    insert("B", 17, saf4,  sSheet ) 

    insert("B", 18, warm5, sSheet )
    insert("B", 19, disc5, sSheet ) 
    insert("B", 20, comp5, sSheet ) 
    insert("B", 21, saf5,  sSheet ) 

    insert("B", 22, warm6, sSheet )
    insert("B", 23, disc6, sSheet ) 
    insert("B", 24, comp6, sSheet ) 
    insert("B", 25, saf6,  sSheet )

    insert("B", 26, warm7, sSheet )
    insert("B", 27, disc7, sSheet )
    insert("B", 28, comp7, sSheet ) 
    insert("B", 29, saf7,  sSheet ) 
 
    insert("B", 30, warm8, sSheet )
    insert("B", 31, disc8, sSheet ) 
    insert("B", 32, comp8, sSheet ) 
    insert("B", 33, saf8,  sSheet )  

    insert("B", 34, warm9, sSheet )
    insert("B", 35, disc9, sSheet ) 
    insert("B", 36, comp9, sSheet )   
    insert("B", 37, saf9,  sSheet )  


    # save into new excel file
    sWorkbook.save(filename="excel-sheets/simplified-arm-study-data.xlsx")


# PRICE WORKBOOK-- UPDATED FROM OWORKBOOKK2, HANDLES PRICE-SENSITIVITY QUESTIONS
#    finds percent difference for each question from what each participant said the kinova arm costs vs what they said the arm they're looking at costs 
def pSheetFunc():

    #copy over names from uSheet first column to aSheet first column
    for i in range (1, (uSheet2.max_row + 1)):
        rows = uSheet2.iter_cols(min_row=i, max_col=1, values_only=True)
        
        #all titles 
        values = [row[0] for row in rows]

        location = "A" + str(i)
        
        pSheet[location] = values[0]


    for i in range(2, (uSheet2.max_column + 1)):
        # how much they said the kinova cost

        kinova = uSheet2.cell(row=2, column=i).value

        if isinstance(kinova, int):

            # copy that value into the top row of data 
            pSheet.cell(row=2, column=i).value = kinova

            #for each arm (located in rows 3 - 11)
            for j in range(3, 12):

                # how much they said this arm costs
                val = int(uSheet2.cell(row=j, column=i).value) 

                # percent diffence between kinova arm price and this arm price 
                pDiff = (val - kinova) / kinova

                pSheet.cell(row=j, column=i).value = pDiff


    # save into new excel file
    pWorkbook.save(filename="excel-sheets/price-sensitivity-arm-study-data.xlsx")


# AVERAGED-PRICE WORKBOOK-- UPDATED FROM PWORKBOOK
#   averages all participant percent differences for each arm and also averages all estimates of kinova arm price
def apSheetFunc():

    #number of columns (represents number of participants)
    num_col = pSheet.max_column

    #copy over names from pSheet first column to apSheet first column
    for i in range (1, (pSheet.max_row + 1)):
        rows = pSheet.iter_cols(min_row=i, max_col=1, values_only=True)
        
        #all titles 
        values = [row[0] for row in rows]

        location = "A" + str(i)
        
        apSheet[location] = values[0]


    #averages all participant responses and copy them into second column (same process in aSheetFunc)
    for i in range (2, (pSheet.max_row + 1)):
        rows = pSheet.iter_cols(min_row=i, min_col=2, max_col=num_col, values_only=True)
        
        #all participant values 
        values = [row[0] for row in rows]

        newValues = []

        #gets rid of 'None' values that prevent doing arithmetic (came as a result of incomplete qualtrics data, shouldn't be a problem anymore idk)
        for j in range(len(values)):

            #some are ints and some are floats but both okay, just don't want 'None' values
            if isinstance(values[j] , int) or isinstance(values[j] , float):
                newValues.append(values[j])

        avg = sum(newValues) / len(newValues)

        location = "B" + str(i)

        apSheet[location] = avg

    #column title
    apSheet["B1"] = "Average"

    # save into new excel file
    apWorkbook.save(filename="excel-sheets/averaged-price-arm-study-data.xlsx")



#SIMPLIFIES THE QUESTION GROUPS AND RETURNS AVERAGE FOR EACH GROUP
#   num is a string (bad naming haha) and represents the arm number
def simplify(num):

    #question numbers distinguish which characteristic category
    #   1-6: Warmth 
    #   7-12: Competency
    #   13-18: Discomfort
    warm = ["1", "2", "3", "4", "5", "6"]
    comp = ["7", "8", "9", "10", "11", "12"]
    disc = ["13", "14", "15", "16", "17", "18"]

    #accumulator variables for each characteristic
    warm1 = 0
    disc1 = 0
    comp1 = 0
    saf1 = 0

    
    #loops for number of rows
    for i in range (1, (aSheet.max_row + 1)):
        rows = aSheet.iter_cols(min_row=i, min_col=1, max_col=2, values_only=True)

        #values in each column of the row
        values = [row[0] for row in rows]

        name = values[0]

        #number of the arm 
        first_letter = str(name)[0]

        #question number
        last_2letter = str(name)[::-1][0:2][::-1]

        #last number 
        last_letter = name[-1]

        #filters for only questions relating to specific arm given as a parameter
        if num in first_letter:

            #rating in cell
            value = values[1]
        
            if isinstance(value, float):

                if "comp" in name: 
                    #warmth questions 
                    if not last_2letter.isnumeric() and last_letter in warm: #or any([x in last_2letter for x in disc]):
                        warm1 += value
                    #competency questions 
                    elif (not last_2letter.isnumeric() and last_letter in comp) or last_2letter in comp:
                        comp1 += value
                    #discomfort questions 
                    else:
                        disc1 += value
                #safety questions
                elif "Average" not in str(value): 
                    saf1 += value
                
    #compute averages
    warmth = warm1/6
    discomfort = disc1/6
    competency = comp1/6
    safety = saf1/3

    return(warmth, discomfort, competency, safety)


#VARIATION OF THE SIMPLIFY FUNCTION ABOVE
#   num is a int that represents the arm number 
#   col is a number that respresents the column number (one person's answers)
def simplify2(num, col, sheet):

    #question numbers distinguish which characteristic category
    #   1-6: Competency
    #   7-12: Discomfort
    warm = ["1", "2", "3", "4", "5", "6"]
    comp = ["7", "8", "9", "10", "11", "12"]
    disc = ["13", "14", "15", "16", "17", "18"]

    #accumulator variables for each characteristic
    warm1 = 0
    disc1 = 0
    comp1 = 0
    saf1 = 0


    for i in range (2, (sheet.max_row + 1)):
    
        #all the rows from value of 'i' to end of data
        rows = sheet.iter_cols(min_row=i, max_col=col, values_only=True)

        #values in each column of the top row (current 'i' value)
        values = [row[0] for row in rows]

        #value in far left column of current row (arm number and char. ex. 1_Warmth)
        name = values[0]

        #number of the arm (ex. 1)
        first_letter = str(name)[0]

        #question number (see breakdown above)
        #   gets last two letters because some question numbers are two digits. For one digit numbers, this will return "_1", "_2", etc. 
        last_2letter = name[::-1][0:2][::-1]

        #last number 
        last_letter = name[-1]

        #if this data pertains to the arm number given by the function parameters
        if str(num) in first_letter:

            value = values[col-1]
        
            # prevents 'None' values from entering this state (which throws errors)
            #    shouldn't be a problem anymore but leaving for increased error checking
            if isinstance(value, int):

                if "comp" in name: 
                    #warmth questions 
                    if not last_2letter.isnumeric() and last_letter in warm: #or any([x in last_2letter for x in disc]):
                        warm1 += value
                    #competency questions 
                    elif (not last_2letter.isnumeric() and last_letter in comp) or last_2letter in comp:
                        comp1 += value
                    #discomfort questions 
                    else:
                        disc1 += value
                #safety questions
                elif "Average" not in str(value): 
                    saf1 += value
            

    #compute averages
    warmth = warm1/6
    discomfort = disc1/6
    competency = comp1/6
    safety = saf1/3

    # error handeling for extra column problem that comes with exporting too many responses
    #   probably will not happen anymore bc of eSheetFunc() but leaving here just incase
    if warmth == 0:
        warmth = None
    if discomfort == 0:
        discomfort = None
    if competency == 0:
        competency = None
    if safety == 0:
        safety = None


    return(warmth, discomfort, competency, safety)



#INSERTS INFO. INTO SPECIFIED LOCATION IN GIVEN SHEET
def insert(col, row, val, sheet):
    loc = col + str(row)
    sheet[loc] = val



#SAVES EXCEL FILES IN CSV FORMAT (for pandas :) )
def csvFunc():
    ## CREATING CSV FILES
    #https://www.studytonight.com/post/converting-xlsx-file-to-csv-file-using-python#:~:text=You%20can%20use%20openpyxl%20to,standard%20file%20I%2FO%20operations.
    uCsv = open("excel-sheets/updated-arm-study-data.csv", "w+")
    qCsv = open("excel-sheets/averaged-questions-arm-study-data.csv", "w+")
    aCsv = open("excel-sheets/averaged-arm-study-data.csv", "w+")
    sCsv = open("excel-sheets/simplified-arm-study-data.csv", "w+")
    pCsv = open("excel-sheets/price-arm-study-data.csv", "w+")

    #for updated sheet
    for row in uSheet.rows:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                uCsv.write(str(l[i].value))
                uCsv.write('\n')
            else:
                uCsv.write(str(l[i].value) + ',')
   
    #for averaged-questions sheet
    for row in qSheet.rows:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                qCsv.write(str(l[i].value))
                qCsv.write('\n')
            else:
                qCsv.write(str(l[i].value) + ',')

    #for averaged sheet
    for row in aSheet.rows:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                aCsv.write(str(l[i].value))
                aCsv.write('\n')
            else:
                aCsv.write(str(l[i].value) + ',')

    #for simplified sheet
    for row in sSheet.rows:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                sCsv.write(str(l[i].value))
                sCsv.write('\n')
            else:
                sCsv.write(str(l[i].value) + ',')

    #for price-column sheet
    for row in pSheet.rows:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                pCsv.write(str(l[i].value))
                pCsv.write('\n')
            else:
                pCsv.write(str(l[i].value) + ',')
            
    uCsv.close()
    qCsv.close()
    aCsv.close()
    sCsv.close()
    sCsv.close()



main()
